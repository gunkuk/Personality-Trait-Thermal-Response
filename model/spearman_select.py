from __future__ import annotations

import os
import numpy as np
import pandas as pd
from dataclasses import dataclass
from typing import Optional, Tuple, List, Dict, Set
from scipy.stats import spearmanr
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

# utils 모듈 임포트 (기존 제공된 utils.py가 같은 경로에 있어야 함)
from utils import (
    norm_columns, first_existing, keep_existing,
    to_numeric_df, lower_triangle_only, reorder_square,
    encode_sex, env_to_set, set_to_direction, make_phase,
    write_excel
)

# =========================
# 1. CONFIGURATION
# =========================
@dataclass(frozen=True)
class Config:
    path_bi: str = r"C:\Users\rjs11\Desktop\1\paper\00_SCI\MBTI\data\BI+SUR.xlsx"
    path_phy_processed: str = r"C:\Users\rjs11\Desktop\1\paper\00_SCI\MBTI\model\OUT\PHY_processed.xlsx"
    out_dir: str = r"C:\Users\rjs11\Desktop\1\paper\00_SCI\MBTI\model\OUT"

    # 분석 대상 그룹 및 순서
    selected_groups: Tuple[str, ...] = ("MBTI", "FFM", "BI", "PHY", "PSY", "BHR", "EUP")
    group_order: Tuple[str, ...] = ("MBTI", "FFM", "BI", "PHY", "PSY", "BHR", "EUP")
    
    @property
    def out_name(self) -> str:
        # 그룹 앞글자 2개씩 따서 파일명 생성
        prefix = "".join([g[:2].upper() for g in self.selected_groups])
        return f"spearman_{prefix}_8_scenarios.xlsx"

    id_col: str = "연번"
    env_col: str = "환경"
    time_col: str = "시간"
    
    # 분석용 시나리오 정의
    scenarios: Tuple[str, ...] = ("A", "B", "C", "D", "E", "F", "G", "H")
    scenario_to_set: Dict[str, str] = None # main에서 초기화

    # 필드 후보군
    sex_candidates: Tuple[str, ...] = ("성별", "sex", "gender")
    age_candidates: Tuple[str, ...] = ("나이", "age")
    bmi_candidates: Tuple[str, ...] = ("bmi",)
    mbti_cols: Tuple[str, ...] = ("e", "n", "t", "j", "a")
    ffm_cols: Tuple[str, ...] = ("o1", "c1", "e1", "a1", "n1")
    resp_cols: Tuple[str, ...] = ("tsv", "tcv", "ta", "tp", "pt")
    p_cols: Tuple[str, ...] = tuple(f"p{i}" for i in range(1, 9))
    m_cols: Tuple[str, ...] = tuple(f"m{i}" for i in range(1, 9))

    # 시간 설정
    static_times: Tuple[int, ...] = (0, 5, 10)
    dynamic_times: Tuple[int, ...] = (10, 15, 20)
    missing_tokens: Tuple[str, ...] = ("-", "—", "–", "", "#n/a", "#na", "n/a", "na")
    alpha: float = 0.05

CFG = Config(scenario_to_set={
    "A": "AB", "B": "AB", "C": "CD", "D": "CD",
    "E": "EF", "F": "EF", "G": "GH", "H": "GH"
})

# =========================
# 2. HELPER FUNCTIONS
# =========================

def rename_to_generic(col_name: str, scenario: str) -> str:
    """
    시나리오별로 다른 접미사를 가진 지표명을 하나의 공통 이름으로 통일합니다.
    예: hr_static_mean -> hr_mean, hr_baseline_1 -> hr_baseline
    """
    s = scenario.upper()
    is_first = s in ("A", "C", "E", "G")
    is_front_half = s in ("A", "B", "C", "D")

    # 1. 시점 접미사 제거
    new_name = col_name.replace("_static", "").replace("_dynamic", "")
    new_name = new_name.replace("_1to10", "").replace("_11to20", "")
    
    # 2. Baseline 번호 통일 (A-D는 1번, E-H는 2번 사용)
    if is_front_half:
        new_name = new_name.replace("_baseline_1", "_baseline")
    else:
        new_name = new_name.replace("_baseline_2", "_baseline")
    
    return new_name

def pairwise_spearman(df: pd.DataFrame, cols: List[str], cfg: Config) -> Tuple[pd.DataFrame, pd.DataFrame]:
    num = to_numeric_df(df, cols, cfg.missing_tokens)
    rho = pd.DataFrame(np.nan, index=cols, columns=cols)
    pval = pd.DataFrame(np.nan, index=cols, columns=cols)
    for i, c1 in enumerate(cols):
        for j, c2 in enumerate(cols[i:], start=i):
            x, y = num[c1], num[c2]
            mask = x.notna() & y.notna()
            if mask.sum() >= 3 and x[mask].nunique() >= 2 and y[mask].nunique() >= 2:
                r, p = spearmanr(x[mask], y[mask])
                rho.loc[c1, c2] = rho.loc[c2, c1] = r
                pval.loc[c1, c2] = pval.loc[c2, c1] = p
    return rho, pval

# =========================
# 3. DATA LOADING & PROCESSING
# =========================

def process_bi_data(cfg: Config) -> Tuple[pd.DataFrame, pd.DataFrame, List[str], List[str]]:
    """Sheet 0 데이터 로드 및 Static/Dynamic 분리 준비"""
    df = norm_columns(pd.read_excel(cfg.path_bi, sheet_name=0))
    id_c = cfg.id_col.lower()

    # 1. 불변 데이터 (성격/데모)
    mbti = keep_existing(df, cfg.mbti_cols)
    ffm = keep_existing(df, cfg.ffm_cols)
    sex_c = first_existing(df, cfg.sex_candidates)
    if sex_c: df["sex_bin"] = encode_sex(df[sex_c])
    age_c = first_existing(df, cfg.age_candidates)
    bmi_c = first_existing(df, cfg.bmi_candidates)
    
    const_cols = mbti + ffm + ["sex_bin"] + ([age_c] if age_c else []) + ([bmi_c] if bmi_c else [])
    bi_const = df[[id_c] + const_cols].drop_duplicates(subset=[id_c])

    # 2. 상태 데이터 (PSY/BHR)
    df["set"] = df[cfg.env_col.lower()].map(env_to_set)
    df["phase"] = make_phase(pd.to_numeric(df[cfg.time_col.lower()], errors="coerce"), cfg.static_times, cfg.dynamic_times)
    
    state_vars = keep_existing(df, cfg.resp_cols + cfg.p_cols + cfg.m_cols)
    phase_means = df.groupby([id_c, "set", "phase"], as_index=False)[state_vars].mean()
    bi_wide = phase_means.pivot(index=[id_c, "set"], columns="phase", values=state_vars)
    bi_wide.columns = [f"{v}_{p}" for v, p in bi_wide.columns]
    
    return bi_const, bi_wide.reset_index(), const_cols, state_vars

def load_sheet2_targets(cfg: Config) -> Tuple[pd.DataFrame, List[str]]:
    df = norm_columns(pd.read_excel(cfg.path_bi, sheet_name=2, header=1))
    id_c = cfg.id_col.lower()
    targets = [c for c in df.columns if c.startswith("eup")] + [c for c in ["sp_c", "sp_h", "dsp"] if c in df.columns]
    return df[[id_c] + targets].copy(), targets

# =========================
# 4. EXCEL STYLING
# =========================
GROUP_FILLS = {
    "MBTI": PatternFill(fgColor="D9EAF7", fill_type="solid"),
    "FFM": PatternFill(fgColor="DFF0D8", fill_type="solid"),
    "BI": PatternFill(fgColor="E6E6E6", fill_type="solid"),
    "PHY": PatternFill(fgColor="FFF2CC", fill_type="solid"),
    "PSY": PatternFill(fgColor="FCE5CD", fill_type="solid"),
    "BHR": PatternFill(fgColor="EADCF8", fill_type="solid"),
    "EUP": PatternFill(fgColor="D9EAD3", fill_type="solid")
}

def apply_styles(out_path: str, sheet_name: str, rho: pd.DataFrame, pval: pd.DataFrame, var_to_group: Dict[str, str], cfg: Config):
    wb = load_workbook(out_path)
    if sheet_name not in wb.sheetnames: return
    ws = wb[sheet_name]
    
    # 헤더 스타일
    for j, c_name in enumerate(rho.columns, start=2):
        if g := var_to_group.get(c_name): ws.cell(1, j).fill = GROUP_FILLS[g]
    for i, r_name in enumerate(rho.index, start=2):
        if g := var_to_group.get(r_name): ws.cell(i, 1).fill = GROUP_FILLS[g]

    # 내용 마스킹 및 색상
    for i, r_name in enumerate(rho.index, start=2):
        for j, c_name in enumerate(rho.columns, start=2):
            val = rho.loc[r_name, c_name]
            if pd.isna(val): continue
            p = pval.loc[r_name, c_name]
            cell = ws.cell(i, j)
            if pd.isna(p) or p >= cfg.alpha:
                cell.fill = PatternFill(fgColor="000000", fill_type="solid")
                cell.font = Font(color="FFFFFF")
            else:
                if val >= 0.6: cell.fill = PatternFill(fgColor="70AD47", fill_type="solid")
                elif val >= 0.4: cell.fill = PatternFill(fgColor="C6E0B4", fill_type="solid")
                elif val <= -0.6: cell.fill = PatternFill(fgColor="E06666", fill_type="solid")
                elif val <= -0.4: cell.fill = PatternFill(fgColor="F4CCCC", fill_type="solid")
    wb.save(out_path)

# =========================
# 5. MAIN PROCESS
# =========================
def main(cfg: Config):
    os.makedirs(cfg.out_dir, exist_ok=True)
    out_path = os.path.join(cfg.out_dir, cfg.out_name)
    id_c = cfg.id_col.lower()

    print("--- 데이터 로딩 중 ---")
    bi_const, bi_wide, const_vars, state_vars = process_bi_data(cfg)
    s2_df, eup_vars = load_sheet2_targets(cfg)
    
    outputs = {}
    pval_storage = {}
    group_map_storage = {}

    for s in cfg.scenarios:
        print(f"시나리오 {s} 분석 중...")
        try:
            # 1. PHY 데이터 로드 (시나리오별 시트)
            phy_df = norm_columns(pd.read_excel(cfg.path_phy_processed, sheet_name=s))
            phy_features = [c for c in phy_df.columns if c not in {id_c, "환경", "시간", "set", "이름"}]
            
            # 2. 매핑 및 통합
            target_set = cfg.scenario_to_set[s]
            is_static = s in ("A", "C", "E", "G")
            suffix = "static" if is_static else "dynamic"

            merged = phy_df.merge(bi_const, on=id_c, how="left")
            merged = merged.merge(bi_wide[bi_wide["set"] == target_set], on=id_c, how="left")
            merged = merged.merge(s2_df, on=id_c, how="left")

            # 3. 지표 이름 통일 및 그룹핑
            final_cols = []
            var_to_group = {}
            
            for g in cfg.group_order:
                g_cols = []
                if g == "PHY": g_cols = phy_features
                elif g in ("MBTI", "FFM", "BI"): g_cols = [c for c in const_vars if c in merged.columns]
                elif g in ("PSY", "BHR"): g_cols = [f"{v}_{suffix}" for v in state_vars if f"{v}_{suffix}" in merged.columns]
                elif g == "EUP": g_cols = [c for c in eup_vars if c in merged.columns]
                
                for c in g_cols:
                    gen_name = rename_to_generic(c, s)
                    merged[gen_name] = merged[c] # 이름 통일된 컬럼 생성
                    if gen_name not in final_cols:
                        final_cols.append(gen_name)
                        var_to_group[gen_name] = g

            # 4. 분석 수행
            rho, pval = pairwise_spearman(merged, final_cols, cfg)
            sheet_name = f"CORR_{s}"
            outputs[sheet_name] = lower_triangle_only(rho, False)
            pval_storage[sheet_name] = pval
            group_map_storage[sheet_name] = var_to_group
            
        except Exception as e:
            print(f"시나리오 {s} 실패: {e}")

    # 결과 저장
    write_excel(out_path, outputs, index=True)
    for sn in outputs.keys():
        apply_styles(out_path, sn, outputs[sn], pval_storage[sn], group_map_storage[sn], cfg)
    
    print(f"--- 모든 공정 완료! 파일 확인: {out_path} ---")

if __name__ == "__main__":
    main(CFG)